import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import google.generativeai as genai
import os
from textwrap import dedent
from typing import Union

# =========================
# è¨­å®šï¼ˆAPIã‚­ãƒ¼ã¯ä»»æ„ï¼‰
# =========================
API_KEY = st.secrets["GEMINI_API_KEY"]  # æœ¬æ¥ãªã‚‰ã°ç’°å¢ƒå¤‰æ•°ã‚’ä½¿ã£ãŸæ–¹ãŒè‰¯ã„ãŒãƒ—ãƒ­ãƒˆã‚¿ã‚¤ãƒ—ã‚¢ãƒ—ãƒªã®ãŸã‚ç›´æ›¸ãã‚’ã—ã¦ã„ã‚‹
if API_KEY:
    genai.configure(api_key=API_KEY)

# =========================
# ãƒ¦ãƒ¼ãƒ†ã‚£ãƒªãƒ†ã‚£
# =========================
WORK_PROCESS_MAP = {
    "1": "èª¿æŸ»åˆ†æã€è¦ä»¶å®šç¾©", "2": "åŸºæœ¬ï¼ˆå¤–éƒ¨ï¼‰è¨­è¨ˆ", "3": "è©³ç´°ï¼ˆå†…éƒ¨ï¼‰è¨­è¨ˆ",
    "4": "ã‚³ãƒ¼ãƒ‡ã‚£ãƒ³ã‚°ãƒ»å˜ä½“ãƒ†ã‚¹ãƒˆ", "5": "ITãƒ»ST", "6": "ã‚·ã‚¹ãƒ†ãƒ é‹ç”¨ãƒ»ä¿å®ˆ",
    "7": "ã‚µãƒ¼ãƒãƒ¼æ§‹ç¯‰ãƒ»é‹ç”¨ç®¡ç†", "8": "DBæ§‹ç¯‰ãƒ»é‹ç”¨ç®¡ç†", "9": "ãƒãƒƒãƒˆãƒ¯ãƒ¼ã‚¯é‹ç”¨ä¿å®ˆ",
    "10": "ãƒ˜ãƒ«ãƒ—ãƒ»ã‚µãƒãƒ¼ãƒˆ", "11": "ãã®ä»–"
}

def safe_str(v) -> str:
    """NaN/NaT/Noneã‚’ç©ºã«ã€æ–‡å­—åˆ—ã¯trimã— 'nan'/'NaT' ã‚‚ç©ºã«ã™ã‚‹"""
    if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "nat"):
        return ""
    return s

def to_str_df(df: pd.DataFrame) -> pd.DataFrame:
    return df.fillna("").astype(str)

def find_first(df_str: pd.DataFrame, keyword: str):
    for r in range(df_str.shape[0]):
        row = df_str.iloc[r]
        for c in range(df_str.shape[1]):
            if keyword in row.iloc[c]:
                return r, c
    return None

def next_right_nonempty(df: pd.DataFrame, r: int, c: int, max_look: int = 20):
    for dc in range(1, max_look + 1):
        cc = c + dc
        if cc >= df.shape[1]:
            break
        v = df.iloc[r, cc]
        s = safe_str(v)
        if s:
            return s
    return ""

def parse_date_like(v) -> Union[date, None]:
    if isinstance(v, (pd.Timestamp, datetime)):
        try:
            return v.date()
        except Exception:
            return None

    if isinstance(v, (int, float)):
        try:                
            # Excelã®ã‚·ãƒªã‚¢ãƒ«å€¤ï¼ˆ1900/1/1ãƒ™ãƒ¼ã‚¹ï¼‰ã¨ã—ã¦å¤‰æ›ã‚’è©¦ã¿ã‚‹
            # '1899-12-30' ã¯Excelã®1900å¹´é–å¹´ãƒã‚°ã‚’è€ƒæ…®ã—ãŸèµ·ç‚¹
            temp_date = pd.to_datetime(v, unit='D', origin='1899-12-30')
            
            return temp_date.date()
        except Exception:
            pass # ã‚·ãƒªã‚¢ãƒ«å€¤ã§ãªã‹ã£ãŸå ´åˆã¯ã€ä¸‹ã®æ–‡å­—åˆ—å‡¦ç†ã¸
    
    s = safe_str(v)
    if not s:
        return None
    # yyyy/mm/dd, yyyy-mm, yyyy/mm, yyyy.mm ã‚’ç·©ãæ‹¾ã†
    # æ—¥ãŒç„¡ã„å ´åˆã¯1æ—¥æ‰±ã„
    m = re.search(r"(\d{4})[./-](\d{1,2})(?:[./-](\d{1,2}))?", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3) or 1)
    try:
        return date(y, mo, d)
    except Exception:
        return None

def looks_like_proc_codes(s: str) -> bool:
    #st.write("ä¸­èº«:", s)
    #st.warning(bool(re.fullmatch(r"[0-9ï¼-ï¼™.ï¼,ï½¤ã€~ã€œï½ ã€€]+", s.strip())))
    return bool(re.fullmatch(r"[0-9ï¼-ï¼™.ï¼,ï½¤ã€~ã€œï½ ã€€]+", s.strip()))

def pick_first_nonempty(values):
    for v in values:
        s = safe_str(v)
        if s:
            return s
    return ""

def uniq_join(lines):
    out = []
    seen = set()
    for x in lines:
        s = safe_str(x)
        if not s:
            continue
        s = s.strip()
        # ç®‡æ¡æ›¸ãè¨˜å·ã®é™¤å»
        if s.startswith(("ãƒ»","-","â€”","â€•","â€“")):
            s = s.lstrip("ãƒ»-â€”â€•â€“ ").strip()
        if s not in seen:
            out.append(s)
            seen.add(s)
    return " / ".join(out)

# =========================
# è§£æ: ã‚·ãƒ¼ãƒˆé¸æŠï¼†èª­ã¿å–ã‚Š
# =========================
LABELS_LEFT = ["ãƒ•ãƒªã‚¬ãƒŠ", "æ°å", "ç¾ä½æ‰€", "æœ€å¯„é§…", "æœ€çµ‚å­¦æ­´"]
LABELS_RIGHT = ["ç”Ÿå¹´æœˆæ—¥", "æ€§åˆ¥", "ç¨¼åƒå¯èƒ½æ—¥"]

def choose_best_sheet(xl: pd.ExcelFile) -> pd.DataFrame:
    best_df, best_score = None, -1
    for sh in xl.sheet_names:
        df = xl.parse(sh, header=None, dtype=object)
        df_str = to_str_df(df)
        score = 0
        for k in LABELS_LEFT + LABELS_RIGHT + ["æƒ…å ±å‡¦ç†è³‡æ ¼", "é …ç•ª", "ä½œæ¥­æœŸé–“", "æ¡ˆä»¶å", "æ¡ˆä»¶åç§°", "ä½œæ¥­å†…å®¹"]:
            if find_first(df_str, k):
                score += 1
        if score > best_score:
            best_df, best_score = df, score
    return best_df

def _collect_rightward_values(df: pd.DataFrame, r: int, c: int, max_cols: int = 12) -> list[str]:
    """è¡Œrã®åˆ—cã®å³å´ã«é€£ç¶šã™ã‚‹éç©ºã‚»ãƒ«ã‚’åé›†ã€‚ç©ºã‚»ãƒ«ã¯ã‚¹ã‚­ãƒƒãƒ—å¯ã ãŒã€å€¤ãŒä¸€åº¦ã‚‚å‡ºãªã„å ´åˆã¯ [] ã‚’è¿”ã™ã€‚"""
    vals = []
    empties_seen = 0
    for dc in range(1, max_cols + 1):
        cc = c + dc
        if cc >= df.shape[1]:
            break
        s = safe_str(df.iloc[r, cc])
        if s:
            vals.append(s)
            empties_seen = 0
        else:
            empties_seen += 1
            # ç©ºã‚»ãƒ«ãŒ2ï½3å€‹é€£ç¶šã—ãŸã‚‰æ‰“ã¡åˆ‡ã‚Šï¼ˆé©åº¦ã«æ—©æœŸçµ‚äº†ï¼‰
            if empties_seen >= 3 and vals:
                break
    return vals

def read_personal(df: pd.DataFrame):
    df_str = to_str_df(df)
    # å·¦å´
    result = {
        "furigana": "", "name": "", "address": "",
        "station": "", "education": "",
        "birth": date(2000,1,1), "gender": "æœªé¸æŠ",
        "available": datetime.now().date(),
        "qualification": ""
    }
    for k in LABELS_LEFT:
        pos = find_first(df_str, k)
        if pos:
            r, c = pos
            result_map = {
                "ãƒ•ãƒªã‚¬ãƒŠ": "furigana",
                "æ°å": "name",
                "ç¾ä½æ‰€": "address",
                "æœ€å¯„é§…": "station",
                "æœ€çµ‚å­¦æ­´": "education",
            }
            result[result_map[k]] = safe_str(next_right_nonempty(df, r, c, 3))
    # å³å´
    # ç”Ÿå¹´æœˆæ—¥
    pos = find_first(df_str, "ç”Ÿå¹´æœˆæ—¥")
    if pos:
        r, c = pos
        b = parse_date_like(next_right_nonempty(df, r, c, 20))
        result["birth"] = b or date(2000,1,1)
    # æ€§åˆ¥
    pos = find_first(df_str, "æ€§åˆ¥")
    if pos:
        r, c = pos
        g = safe_str(next_right_nonempty(df, r, c, 20))
        if g in ["ç”·", "ç”·æ€§"]:
            result["gender"] = "ç”·æ€§"
        elif g in ["å¥³", "å¥³æ€§"]:
            result["gender"] = "å¥³æ€§"
        elif g == "ãã®ä»–":
            result["gender"] = "ãã®ä»–"
        else:
            result["gender"] = "æœªé¸æŠ"
    # ç¨¼åƒå¯èƒ½æ—¥
    pos = find_first(df_str, "ç¨¼åƒå¯èƒ½æ—¥")
    if pos:
        r, c = pos
        s = safe_str(next_right_nonempty(df, r, c, 20))
        if ("å³æ—¥" in s) or (s in ["-", "--", ""]):
            result["available"] = datetime.now().date()
        else:
            d = parse_date_like(s)
            result["available"] = d or datetime.now().date()
    # è³‡æ ¼ï¼ˆè¡Œå†…ã®å³æ–¹å‘ã‚’ã™ã¹ã¦åé›†ã€‚ãªã‘ã‚Œã°æ•°è¡Œä¸‹ã‚‚ã‚¹ã‚­ãƒ£ãƒ³ï¼‰
    pos = find_first(df_str, "æƒ…å ±å‡¦ç†è³‡æ ¼")
    if pos:
        r, c = pos
        vals = _collect_rightward_values(df, r, c, max_cols=12)
        if not vals:
            # è¡Œå†…ã«è¦‹ã¤ã‹ã‚‰ãªã„å ´åˆã¯ã€ä¸‹æ–¹å‘ï¼ˆæ¬¡ã®5è¡Œï¼‰ã§å³å´ã®å€¤ã‚’æ¢ç´¢
            for rr in range(r+1, min(r+6, df.shape[0])):
                vals.extend(_collect_rightward_values(df, rr, c, max_cols=12))
        result["qualification"] = "\n".join([safe_str(v) for v in vals if safe_str(v)])
    return result

def find_header_row(df_str: pd.DataFrame) -> Union[int, None]:
    for r in range(df_str.shape[0]):
        row_vals = [df_str.iloc[r, c] for c in range(df_str.shape[1])]
        cond1 = any("é …" in v for v in row_vals) and any("ä½œæ¥­æœŸé–“" in v for v in row_vals)
        cond2 = any(("æ¡ˆä»¶å" in v) or ("æ¡ˆä»¶åç§°" in v) for v in row_vals)
        cond3 = any("ä½œæ¥­å†…å®¹" in v for v in row_vals)
        if cond1 and cond2 and cond3:
            return r
    return None

def col_at(df_str: pd.DataFrame, r: int, keywords: list[str]) -> Union[int, None]:
    for c in range(df_str.shape[1]):
        cell = df_str.iloc[r, c]
        if any(k in cell for k in keywords):
            return c
    return None

def col_at_multiheader(df_str: pd.DataFrame, rows: list[int], keywords: list[str]) -> Union[int, None]:
    """ãƒ˜ãƒƒãƒ€è¡Œã¨ã‚µãƒ–ãƒ˜ãƒƒãƒ€è¡Œã®ä¸¡æ–¹ã‚’è¦‹ã¦æœ€åˆã«ä¸€è‡´ã—ãŸåˆ—ã‚’è¿”ã™"""
    for r in rows:
        c = col_at(df_str, r, keywords)
        if c is not None:
            return c
    return None

def parse_projects(df: pd.DataFrame) -> list:
    df_str = to_str_df(df)
    header_r = find_header_row(df_str)
    if header_r is None:
        return []

    subheader_r = header_r + 1 if header_r + 1 < df.shape[0] else header_r

    # åˆ—ä½ç½®ã®æ¨å®šï¼ˆãƒ˜ãƒƒãƒ€è¡Œï¼‹ã‚µãƒ–ãƒ˜ãƒƒãƒ€è¡Œã‚’è€ƒæ…®ï¼‰
    C_ID = col_at_multiheader(df_str, [header_r, subheader_r], ["é …ç•ª", "é …"])
    C_PERIOD = col_at_multiheader(df_str, [header_r, subheader_r], ["ä½œæ¥­æœŸé–“", "æœŸé–“"])
    C_NAME = col_at_multiheader(df_str, [header_r, subheader_r], ["æ¡ˆä»¶å", "æ¡ˆä»¶åç§°", "æ¡ˆä»¶"])
    C_CONTENT = col_at_multiheader(df_str, [header_r, subheader_r], ["ä½œæ¥­å†…å®¹", "å†…å®¹"])
    C_OS = col_at_multiheader(df_str, [header_r, subheader_r], ["OS"])
    C_LANG = col_at_multiheader(df_str, [header_r, subheader_r], ["è¨€èª", "ãƒ„ãƒ¼ãƒ«"])
    C_DB = col_at_multiheader(df_str, [header_r, subheader_r], ["DB", "DB/DC", "DC"])
    C_PROC = col_at_multiheader(df_str, [header_r, subheader_r], ["ä½œæ¥­å·¥ç¨‹", "å·¥ç¨‹"])
    C_ROLE = col_at_multiheader(df_str, [header_r, subheader_r], ["å½¹å‰²"])
    C_POS = col_at_multiheader(df_str, [header_r, subheader_r], ["ãƒã‚¸ã‚·ãƒ§ãƒ³", "å½¹è·"])
    C_SCALE = col_at_multiheader(df_str, [header_r, subheader_r], ["è¦æ¨¡", "äººæ•°"])

    # å¿…é ˆ
    if any(x is None for x in [C_PERIOD, C_NAME, C_CONTENT]):
        return []

    projects = []
    cur = None

    def cell(r, cidx):
        return safe_str(df.iloc[r, cidx]) if (cidx is not None and cidx < df.shape[1]) else ""

    def flush_cur():
        nonlocal cur
        if not cur:
            return
        # æœŸé–“â†’é–‹å§‹/çµ‚äº†æ¨å®š
        dates = []
        for s in cur["periods"]:
            d = parse_date_like(s)
            if d:
                dates.append(d)
        start_date = min(dates) if dates else None
        end_date = max(dates) if dates else None
        # ã€Œç¾ã€ã€Œç¾åœ¨ã€å¯¾ç­–
        txt_all = " ".join(cur["periods"])
        if re.search(r"(ç¾|ç¾åœ¨)", txt_all):
            end_date = datetime.now().date()

        # ä½œæ¥­å·¥ç¨‹ï¼ˆç•ªå·â†’ãƒ©ãƒ™ãƒ«ï¼‰
        proc_labels = []
        for s in cur["procs"]:
            s_raw = s.strip()
            if looks_like_proc_codes(s_raw):
                s_normalized = s_raw.translate(str.maketrans({
                    # å…¨è§’æ•°å­— -> åŠè§’æ•°å­—
                    'ï¼': '0', 'ï¼‘': '1', 'ï¼’': '2', 'ï¼“': '3', 'ï¼”': '4',
                    'ï¼•': '5', 'ï¼–': '6', 'ï¼—': '7', 'ï¼˜': '8', 'ï¼™': '9',
                    # å…¨è§’è¨˜å· -> åŠè§’ã¾ãŸã¯çµ±ä¸€è¨˜å·
                    'ï¼': '.',  # å…¨è§’ãƒ‰ãƒƒãƒˆ -> åŠè§’ãƒ‰ãƒƒãƒˆ
                    'ï½¤': ',',  # å…¨è§’ã‚«ãƒ³ãƒ -> åŠè§’ã‚«ãƒ³ãƒ
                    'ã€': ',',  # èª­ç‚¹ -> åŠè§’ã‚«ãƒ³ãƒ
                    'ï¼Œ': ',',   # å…¨è§’ã‚«ãƒ³ãƒ(FF0C) -> åŠè§’ã‚«ãƒ³ãƒ
                    'ï½': 'ã€œ',  # å…¨è§’ãƒãƒ«ãƒ€ -> æ³¢ãƒ€ãƒƒã‚·ãƒ¥(ç¯„å›²è¨˜å·ã¨ã—ã¦çµ±ä¸€)
                    '~': 'ã€œ',  # åŠè§’ãƒãƒ«ãƒ€ -> æ³¢ãƒ€ãƒƒã‚·ãƒ¥
                }))

                final_codes = [] 
                
                parts = re.split(r"[.,]+", s_normalized)
                for part in parts:
                    part = part.strip()
                    if not part:
                        continue
                    
                    range_match = re.search(r"^(\d+)\s*ã€œ\s*(\d+)$", part) 
                    
                    if range_match:
                        try:
                            start = int(range_match.group(1)) 
                            end = int(range_match.group(2))   
                            for i in range(start, end + 1):
                                final_codes.append(str(i))
                        except ValueError:
                            pass 
                    else:
                        # ç¯„å›²ã§ãªã„å ´åˆï¼ˆå˜ãªã‚‹æ•°å­—ï¼‰
                        if re.fullmatch(r"\d+", part):
                            final_codes.append(part)
                
                for k in [x for x in final_codes if x]:
                    if k in WORK_PROCESS_MAP and WORK_PROCESS_MAP[k] not in proc_labels:
                        proc_labels.append(WORK_PROCESS_MAP[k])
                             
                #for k in [x for x in re.split(r"[.,]+", s2) if x]:
                #    if k in WORK_PROCESS_MAP and WORK_PROCESS_MAP[k] not in proc_labels:
                #        proc_labels.append(WORK_PROCESS_MAP[k])
            else:
                for key, label_name in WORK_PROCESS_MAP.items():
                    if label_name == s_raw:
                        proc_labels.append(label_name)
    

        projects.append({
            "start_date": start_date or date(2000,1,1),
            "end_date": end_date or datetime.now().date(),
            "project_name": cur["project_name"] or "",
            "industry": cur["industry"] or "",
            "work_content": "\n".join([s for s in cur["contents"] if s]),
            "os": uniq_join(cur["oss"]),
            "db_dc": uniq_join(cur["dbs"]),
            "lang_tool": uniq_join(cur["langs"]),
            "work_process_list": proc_labels,
            "work_process_str": ", ".join(proc_labels),
            "role": cur["role"] or "",
            "position": cur["position"] or "",
            "scale": cur["scale"] or "",
        })

    for r in range(subheader_r + 1, df.shape[0]):
        idv = cell(r, C_ID)
        is_new = bool(re.search(r"\d", idv))  # æ•°å­—ãŒå…¥ã£ã¦ã„ã‚Œã°æ–°æ¡ˆä»¶

        if is_new:
            if cur:
                flush_cur()
            cur = {
                "project_name": None,
                "industry": None,
                "periods": [],
                "contents": [],
                "oss": [],
                "dbs": [],
                "langs": [],
                "procs": [],
                "role": "",
                "position": "",
                "scale": "",
            }

        if not cur:
            continue  # ã¾ã ãƒ˜ãƒƒãƒ€ç›´ä¸‹ã®ç©ºè¡Œãªã©

        # åŸºæœ¬ã‚»ãƒ«
        period_val = cell(r, C_PERIOD)
        cur["periods"].append(period_val)
        is_firstline = bool(parse_date_like(period_val))  # æ¡ˆä»¶1è¡Œç›®ã‹ã©ã†ã‹

        name_val = cell(r, C_NAME)
        if name_val:
            if is_firstline and cur["project_name"] is None:
                cur["project_name"] = name_val
            elif (not is_firstline) and cur["industry"] is None:
                # ã€Œæ¡ˆä»¶åç§°ã®çœŸä¸‹ã€ã‚’æ¥­ç¨®ã¨ã—ã¦æ‹¾ã†
                cur["industry"] = name_val
            elif cur["project_name"] is None:
                cur["project_name"] = name_val  # ãƒ•ã‚©ãƒ¼ãƒ«ãƒãƒƒã‚¯

        content_val = cell(r, C_CONTENT)
        if content_val:
            cur["contents"].append(content_val)

        os_val = cell(r, C_OS)
        if os_val:
            cur["oss"].append(os_val)

        lang_val = cell(r, C_LANG)
        if lang_val:
            for t in re.split(r"[ã€,/\n]+", lang_val):
                t = t.strip().lstrip("-ãƒ»").strip()
                if t:
                    cur["langs"].append(t)

        db_val = cell(r, C_DB)
        if db_val:
            for t in re.split(r"[ã€,/\n]+", db_val):
                t = t.strip().lstrip("-ãƒ»").strip()
                if t:
                    cur["dbs"].append(t)

        proc_val = cell(r, C_PROC)
        target_initials = ("èª¿æŸ»åˆ†æã€è¦ä»¶å®šç¾©", "åŸºæœ¬ï¼ˆå¤–éƒ¨ï¼‰è¨­è¨ˆ", "è©³ç´°ï¼ˆå†…éƒ¨ï¼‰è¨­è¨ˆ",
                            "ã‚³ãƒ¼ãƒ‡ã‚£ãƒ³ã‚°ãƒ»å˜ä½“ãƒ†ã‚¹ãƒˆ", "ITãƒ»ST", "ã‚·ã‚¹ãƒ†ãƒ é‹ç”¨ãƒ»ä¿å®ˆ",
                            "ã‚µãƒ¼ãƒãƒ¼æ§‹ç¯‰ãƒ»é‹ç”¨ç®¡ç†", "DBæ§‹ç¯‰ãƒ»é‹ç”¨ç®¡ç†", "ãƒãƒƒãƒˆãƒ¯ãƒ¼ã‚¯é‹ç”¨ä¿å®ˆ",
                            "ãƒ˜ãƒ«ãƒ—ãƒ»ã‚µãƒãƒ¼ãƒˆ", "ãã®ä»–",
                            "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11",
                            "ï¼‘", "ï¼’", "ï¼“", "ï¼”", "ï¼•", "ï¼–", "ï¼—", "ï¼˜", "ï¼™", "ï¼‘ï¼", "ï¼‘ï¼‘")
        #if proc_val and is_firstline:
        if proc_val and proc_val.startswith(target_initials):
            #st.write(r, proc_val)
            cur["procs"].append(proc_val)

        role_val = cell(r, C_ROLE)
        target_initials = ("S", "P")
        if role_val and role_val.startswith(target_initials):
            cur["role"] = role_val

        pos_val = cell(r, C_POS)
        if pos_val and (not is_firstline) and not cur["position"]:
            cur["position"] = pos_val

        scale_val = cell(r, C_SCALE)
        if scale_val and is_firstline and not cur["scale"]:
            cur["scale"] = scale_val

    if cur:
        flush_cur()

    # ç©ºæ¡ˆä»¶ã‚’é™¤å»ï¼ˆåç§°ã‚‚å†…å®¹ã‚‚ç©ºï¼‰
    projects = [p for p in projects if (p["project_name"] or p["work_content"])]
    return projects

# =========================
# Session åˆæœŸåŒ–
# =========================
def initialize_session_state():
    ss = st.session_state
    ss.setdefault("pi_name", "")
    ss.setdefault("pi_furigana", "")
    ss.setdefault("pi_birth_date", date(2000,1,1))
    ss.setdefault("pi_gender", "æœªé¸æŠ")
    ss.setdefault("pi_address", "")
    ss.setdefault("pi_nearest_station", "")
    ss.setdefault("pi_education", "")
    ss.setdefault("pi_available_date", datetime.now().date())
    ss.setdefault("pi_qualifications_input", "")
    ss.setdefault("pi_summary", "")
    ss.setdefault("projects", [])
    ss.setdefault("generated_overview", "")

initialize_session_state()

# =========================
# ã‚³ãƒ¼ãƒ«ãƒãƒƒã‚¯
# =========================
def load_from_excel_callback():
    uploaded_file = st.session_state.excel_uploader
    if uploaded_file is None:
        return
    try:
        xl = pd.ExcelFile(uploaded_file)
        df = choose_best_sheet(xl)
        if df is None:
            st.error("æœ‰åŠ¹ãªã‚·ãƒ¼ãƒˆãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚")
            return

        # --- å€‹äººæƒ…å ±ï¼†è³‡æ ¼ ---
        pi = read_personal(df)
        st.session_state.pi_furigana = pi["furigana"]
        st.session_state.pi_name = pi["name"]
        st.session_state.pi_address = pi["address"]
        st.session_state.pi_nearest_station = pi["station"]
        st.session_state.pi_education = pi["education"]
        st.session_state.pi_birth_date = pi["birth"]
        st.session_state.pi_gender = pi["gender"]
        st.session_state.pi_available_date = pi["available"]
        st.session_state.pi_qualifications_input = pi["qualification"]

        # --- æ¥­å‹™çµŒæ­´ ---
        st.session_state.projects = parse_projects(df)

        st.success("Excelã®å†…å®¹ã‚’å…¥åŠ›æ¬„ã¸åæ˜ ã—ã¾ã—ãŸã€‚")

    except Exception as e:
        st.error(f"èª­ã¿è¾¼ã¿ä¸­ã«ã‚¨ãƒ©ãƒ¼: {e}")

def enhance_with_ai_callback():
    if not API_KEY:
        st.warning("Gemini APIã‚­ãƒ¼ãŒæœªè¨­å®šã®ãŸã‚ã‚¹ã‚­ãƒƒãƒ—ã—ã¾ã—ãŸã€‚")
        return
    try:
        model = genai.GenerativeModel("gemini-2.5-flash")
        # ã‚µãƒãƒª
        prompt1 = dedent("""
            ã‚ãªãŸã¯çµŒé¨“è±Šå¯Œãªã‚­ãƒ£ãƒªã‚¢ã‚¢ãƒ‰ãƒã‚¤ã‚¶ãƒ¼ã§ã™ã€‚ä»¥ä¸‹ã®ã€Œé–‹ç™ºçµŒé¨“ã‚µãƒãƒªã€ã‚’ã€
            ç°¡æ½”ã§å°‚é–€çš„ãªè¡¨ç¾ã«æ•´ãˆã¦ãã ã•ã„ã€‚å‡ºåŠ›ã¯ä¿®æ­£å¾Œã®æœ¬æ–‡ã®ã¿ã€‚
        """) + "\n" + st.session_state.pi_summary
        st.session_state.pi_summary = model.generate_content(prompt1).text

        # å„æ¡ˆä»¶
        for i, p in enumerate(st.session_state.projects):
            if p.get("work_content"):
                prompt2 = dedent("""
                    ã‚ãªãŸã¯çµŒé¨“è±Šå¯Œãªã‚­ãƒ£ãƒªã‚¢ã‚¢ãƒ‰ãƒã‚¤ã‚¶ãƒ¼ã§ã™ã€‚ä»¥ä¸‹ã®ã€Œä½œæ¥­å†…å®¹ã€ã‚’ã€
                    å®Ÿç¸¾ãŒç°¡æ½”ã«ä¼ã‚ã‚‹ã‚ˆã†ã«ç®‡æ¡æ›¸ãã«æ•´ãˆã¦ãã ã•ã„ã€‚å‡ºåŠ›ã¯æœ¬æ–‡ã®ã¿ã€‚
                """) + "\n" + p["work_content"]
                st.session_state.projects[i]["work_content"] = model.generate_content(prompt2).text
        st.success("AIã§æ–‡ç« ã‚’æ•´å½¢ã—ã¾ã—ãŸã€‚")
    except Exception as e:
        st.error(f"AIå‡¦ç†ã§ã‚¨ãƒ©ãƒ¼: {e}")

def generate_overview_callback():
    try:
        skills = set()
        for p in st.session_state.projects:
            if p.get("os"): skills.update([s.strip() for s in str(p["os"]).split("/") if s.strip()])
            if p.get("db_dc"): skills.update([s.strip() for s in str(p["db_dc"]).split("/") if s.strip()])
            if p.get("lang_tool"): skills.update([s.strip() for s in str(p["lang_tool"]).split("/") if s.strip()])
            if p.get("work_process_list"): skills.update(p["work_process_list"])
        all_work = "\n".join([str(p.get("work_content","")) for p in st.session_state.projects])
        remarks = ""
        if API_KEY:
            model = genai.GenerativeModel("gemini-2.5-flash")
            prompt = dedent("""
                ä»¥ä¸‹ã®ä½œæ¥­å†…å®¹ã‚’è¦ç´„ã—ã€å‚™è€ƒã¨ã—ã¦1ï½2æ–‡ã§å‡ºåŠ›ã—ã¦ãã ã•ã„ã€‚å‡ºåŠ›ã¯æœ¬æ–‡ã®ã¿ã€‚
            """) + "\n" + all_work
            remarks = model.generate_content(prompt).text
        age = (datetime.now().date() - st.session_state.pi_birth_date).days // 365
        gender_str = {"æœªé¸æŠ":"", "ç”·æ€§":"ç”·", "å¥³æ€§":"å¥³", "ãã®ä»–":""}.get(st.session_state.pi_gender, "")

        lines = [
            f"æ°å\t:{st.session_state.pi_name}ã€€{age}æ­³ã€€{gender_str}",
            f"æœ€å¯„\t:{st.session_state.pi_nearest_station}",
            "é–‹å§‹\t:å³æ—¥å¯ï½",
            "å˜ä¾¡\t:",
            f"ã‚¹ã‚­ãƒ«\t:{', '.join(sorted(list(skills)))}",
            f"è³‡æ ¼\t:{st.session_state.pi_qualifications_input.replace(chr(10), ', ')}",
            f"å‚™è€ƒ\t:{remarks}"
        ]
        overview_text = "\n".join(lines)

        st.session_state.generated_overview = overview_text.strip()
        st.success("æ¦‚è¦ã‚’ä½œæˆã—ã¾ã—ãŸã€‚")
    except Exception as e:
        st.error(f"æ¦‚è¦ä½œæˆã‚¨ãƒ©ãƒ¼: {e}")

# =========================
# UI
# =========================
st.set_page_config(page_title="ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆè‡ªå‹•å…¥åŠ›ï¼†Geminiè¦ç´„ã‚¢ãƒ—ãƒª", layout="centered")
st.title("ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆè‡ªå‹•å…¥åŠ›ï¼†Geminiè¦ç´„ã‚¢ãƒ—ãƒª")
st.caption("çµŒæ­´æ›¸Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ã—ã¦ãã ã•ã„")

with st.sidebar:
    st.header("ğŸ“‚ ã‚µã‚¤ãƒ‰ãƒ¡ãƒ‹ãƒ¥ãƒ¼")
    page = st.radio("ãƒšãƒ¼ã‚¸é¸æŠ", ["ãƒ›ãƒ¼ãƒ ", "åŸºæœ¬æƒ…å ±", "é–‹ç™ºçµŒé¨“ã‚µãƒãƒª", "æ¥­å‹™å±¥æ­´", "AIã«ã‚ˆã‚‹æ”¹å–„"])
    
uploaded_file = st.file_uploader(
    "Excelãƒ•ã‚¡ã‚¤ãƒ«ï¼ˆ.xlsxæ¨å¥¨ï¼‰",
    type=["xlsx", "csv"],
    key="excel_uploader",
    on_change=load_from_excel_callback
)

def basic_info():
    st.header("å€‹äººæƒ…å ±")
    cols = st.columns(2)
    with cols[0]:
        st.session_state.pi_furigana = st.text_input("ãƒ•ãƒªã‚¬ãƒŠ", st.session_state.pi_furigana)
        st.session_state.pi_name = st.text_input("æ°å", st.session_state.pi_name)
        st.session_state.pi_address = st.text_input("ç¾ä½æ‰€", st.session_state.pi_address)
        st.session_state.pi_nearest_station = st.text_input("æœ€å¯„é§…", st.session_state.pi_nearest_station)
    with cols[1]:
        st.session_state.pi_birth_date = st.date_input("ç”Ÿå¹´æœˆæ—¥", st.session_state.pi_birth_date)
        st.session_state.pi_gender = st.selectbox("æ€§åˆ¥", ["æœªé¸æŠ","ç”·æ€§","å¥³æ€§","ãã®ä»–"], index=["æœªé¸æŠ","ç”·æ€§","å¥³æ€§","ãã®ä»–"].index(st.session_state.pi_gender))
        st.session_state.pi_available_date = st.date_input("ç¨¼åƒå¯èƒ½æ—¥", st.session_state.pi_available_date)
        st.session_state.pi_education = st.text_input("æœ€çµ‚å­¦æ­´", st.session_state.pi_education)

    st.subheader("æƒ…å ±å‡¦ç†è³‡æ ¼")
    st.session_state.pi_qualifications_input = st.text_area("è‡ªç”±è¨˜è¿°", value=st.session_state.pi_qualifications_input, height=100)

def deve_expe():
    st.subheader("é–‹ç™ºçµŒé¨“ã‚µãƒãƒª")
    st.session_state.pi_summary = st.text_area("è‡ªç”±è¨˜è¿°", value=st.session_state.pi_summary, 
                                  placeholder="ä¾‹ï¼š\nçµŒé¨“å¹´æ•°\nä½¿ç”¨ã—ã¦ããŸæŠ€è¡“(è¨€èªã€ãƒ•ãƒ¬ãƒ¼ãƒ ãƒ¯ãƒ¼ã‚¯ã€ã‚¯ãƒ©ã‚¦ãƒ‰ãªã©)\nå½¹å‰²(è¦ä»¶å®šç¾©ã€åŸºæœ¬è¨­è¨ˆã€å®Ÿè£…ã€é‹ç”¨ãªã©)\nå®Ÿç¸¾ãƒ»å¾—æ„åˆ†é‡", height=160)

def business_history():
    st.header("æ¥­å‹™çµŒæ­´")
    if st.button("æ–°ã—ã„æ¡ˆä»¶ã‚’è¿½åŠ "):
        st.session_state.projects.append({})

    roles = ["PM", "PL", "SPL", "SE", "PG"]
    roles_with_name = ["PM ãƒ—ãƒ­ã‚¸ã‚§ã‚¯ãƒˆãƒãƒãƒ¼ã‚¸ãƒ£ãƒ¼", "PL ãƒ—ãƒ­ã‚¸ã‚§ã‚¯ãƒˆãƒªãƒ¼ãƒ€ãƒ¼",
                   "SPL ã‚µãƒ–ãƒªãƒ¼ãƒ€ãƒ¼", "SE ã‚·ã‚¹ãƒ†ãƒ ã‚¨ãƒ³ã‚¸ãƒ‹ã‚¢", "PG ãƒ—ãƒ­ã‚°ãƒ©ãƒãƒ¼"]

    for i, p in enumerate(st.session_state.projects):
        st.subheader(f"æ¡ˆä»¶ {i+1}")
        cols = st.columns(2)

        if p.get("role", "") in roles:
            idx = roles.index(p.get("role", ""))
        elif p.get("role", "") in roles_with_name:
            idx = roles_with_name.index(p.get("role", ""))
        else:
            idx = None
    
        with cols[0]:
            p["start_date"] = st.date_input(f"é–‹å§‹æ—¥ (æ¡ˆä»¶ {i+1})", p.get("start_date", date(2022,4,1)))
            p["end_date"] = st.date_input(f"çµ‚äº†æ—¥ (æ¡ˆä»¶ {i+1})", p.get("end_date", datetime.now().date()))
            p["project_name"] = st.text_input(f"æ¡ˆä»¶åç§° (æ¡ˆä»¶ {i+1})", p.get("project_name",""))
            p["industry"] = st.text_input(f"æ¥­ç¨® (æ¡ˆä»¶ {i+1})", p.get("industry",""))
        with cols[1]:
            p["os"] = st.text_input(f"OS (æ¡ˆä»¶ {i+1})", p.get("os",""))
            p["db_dc"] = st.text_input(f"DB/DC (æ¡ˆä»¶ {i+1})", p.get("db_dc",""))
            p["lang_tool"] = st.text_input(f"è¨€èª/ãƒ„ãƒ¼ãƒ« (æ¡ˆä»¶ {i+1})", p.get("lang_tool",""))
            p["role"] = st.selectbox(f"å½¹å‰² (æ¡ˆä»¶ {i+1})", roles_with_name, index=idx)
            #p["role"] = st.text_input(f"å½¹å‰² (æ¡ˆä»¶ {i+1})", p.get("role",""))
            p["position"] = st.text_input(f"ãƒã‚¸ã‚·ãƒ§ãƒ³ (æ¡ˆä»¶ {i+1})", p.get("position",""))
            p["scale"] = st.text_input(f"è¦æ¨¡ (æ¡ˆä»¶ {i+1})", p.get("scale",""))
        p["work_content"] = st.text_area(f"ä½œæ¥­å†…å®¹ (æ¡ˆä»¶ {i+1})", p.get("work_content",""))
        selected = st.multiselect(
            f"ä½œæ¥­å·¥ç¨‹ (æ¡ˆä»¶ {i+1})",
            options=list(WORK_PROCESS_MAP.keys()),
            format_func=lambda k: WORK_PROCESS_MAP[k],
            default=[k for k, v in WORK_PROCESS_MAP.items() if v in p.get("work_process_list", [])]
        )
        p["work_process_list"] = [WORK_PROCESS_MAP[k] for k in selected]
        p["work_process_str"] = ", ".join(p["work_process_list"])
        if st.button(f"ã“ã®æ¡ˆä»¶ã‚’å‰Šé™¤ (æ¡ˆä»¶ {i+1})"):
            st.session_state.projects.pop(i)
            st.rerun()
        st.markdown("---")

def ai_impr():
    st.header("ç”ŸæˆAIã«ã‚ˆã‚‹ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆæ”¹å–„")
    st.button("ç”ŸæˆAIã«æ”¹å–„ã‚’ä¾é ¼", on_click=enhance_with_ai_callback)

    st.header("ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆæ¦‚è¦ã®æŠ½å‡º")
    st.button("æ¦‚è¦ã‚’æŠ½å‡º", on_click=generate_overview_callback)
    if st.session_state.generated_overview:
        st.text_area("æŠ½å‡ºã•ã‚ŒãŸæ¦‚è¦", value=st.session_state.generated_overview, height=240)

    # ---- Excelå‡ºåŠ›ï¼ˆæ·»ä»˜ãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼ã«ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆå¤‰æ›´ï¼‰ ----
    if st.button("ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆã‚’ç”Ÿæˆ (Excelå½¢å¼)"):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            if "Sheet" in wb.sheetnames:
                 wb.remove(wb["Sheet"])
            ws = wb.create_sheet("ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆ")
            wb.active = ws

            # --- ã‚¹ã‚¿ã‚¤ãƒ«å®šç¾© ---
            # (æ·»ä»˜ãƒ•ã‚¡ã‚¤ãƒ«å½¢å¼ã«åˆã‚ã›ã¦ã€èƒŒæ™¯è‰²ãªã©ã‚’èª¿æ•´)
            title_font = Font(size=24, bold=True)
            section_title_font = Font(bold=True, size=12) # èƒŒæ™¯è‰²ãªã—
            work_history_font = Font(bold=False, size=9) # èƒŒæ™¯è‰²ãªã—
            numbering_font = Font(bold=False, size=8)
            bold_font = Font(bold=True, size=10)
            data_font = Font(bold=False, size=10)
        
            # æ¥­å‹™çµŒæ­´ãƒ†ãƒ¼ãƒ–ãƒ«ãƒ˜ãƒƒãƒ€ç”¨ã®èƒŒæ™¯è‰²
            project_title_fill = PatternFill(fgColor="D3D3D3", fill_type="solid")
        
            # ç½«ç·š
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            dashdot_border = Border(left=Side(style='dashDot'), right=Side(style='dashDot'), top=None, bottom=None)
            data_border = Border(left=Side(style='thin'), right=Side(style='dashDot'), top=None, bottom=None)
        
            # æŠ˜ã‚Šè¿”ã— + ä¸Šå¯„ã›
            wrap_text_alignment = Alignment(wrapText=True, vertical='top')
            center_text_alignment = Alignment(horizontal = 'center', vertical = 'center')

            # ãƒ†ãƒ¼ãƒ–ãƒ«ã®åˆ—æ•°ï¼ˆKåˆ—ã¾ã§ï¼‰
            TABLE_COLS = 11
            # ä½œæ¥­å†…å®¹ã‚’æ›¸ãè¾¼ã‚€åˆ— (Cåˆ—)
            COL_PROJECT_NAME = 5

            cur = 1 # ç¾åœ¨ã®è¡Œç•ªå·

            # --- 1è¡Œç›®: ç©ºç™½ ---
            cur += 1 # 2è¡Œç›®ã‹ã‚‰ã‚¹ã‚¿ãƒ¼ãƒˆ

            # --- ãƒ˜ãƒ«ãƒ‘ãƒ¼é–¢æ•° ---
            def style(cell, font=None, fill=None, border=None, align=None):
                if font: cell.font = font
                if fill: cell.fill = fill
                if border: cell.border = border
                if align: cell.alignment = align

            # --- 2è¡Œç›®: ã‚¿ã‚¤ãƒˆãƒ« ---
            cell = ws.cell(row=cur, column=2, value="æ¥­å‹™çµŒæ­´æ›¸")
            style(cell, font=title_font, align=center_text_alignment, border=thin_border)
            ws.merge_cells('B2:K3')
            cur += 2 # 3è¡Œç›®ã¯ç©ºç™½ã€4è¡Œç›®ã‹ã‚‰
        
            # --- 4è¡Œç›®: 1. å€‹äººæƒ…å ± ---
            rows = [
                ("ãƒ•ãƒªã‚¬ãƒŠ", st.session_state.pi_furigana, "ç”Ÿå¹´æœˆæ—¥", st.session_state.pi_birth_date.strftime("%Y/%m/%d")),
                ("æ°å", st.session_state.pi_name, "æ€§åˆ¥", st.session_state.pi_gender),
                ("ç¾ä½æ‰€", st.session_state.pi_address, "ç¨¼åƒå¯èƒ½æ—¥", st.session_state.pi_available_date.strftime("%Y/%m/%d")),
                ("æœ€å¯„é§…", st.session_state.pi_nearest_station),
                ("æœ€çµ‚å­¦æ­´", st.session_state.pi_education, ),
            ]
            count = 0
            for row in rows:
                style(ws.cell(row=cur, column=2, value=row[0]), font=bold_font, border=thin_border)
                style(ws.cell(row=cur, column=4, value=row[1]), border=thin_border)

                if len(row) == 4 :
                    style(ws.cell(row=cur, column=7, value=row[2]), font=bold_font, border=thin_border)
                
                    if count == 0:
                        style(ws.cell(row=cur, column=9, value=row[3]), border=thin_border)
                        count = 1
                    else:
                        style(ws.cell(row=cur, column=10, value=row[3]), border=thin_border)
                cur += 1

            # ãƒ•ãƒªã‚¬ãƒŠ
            ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3)
            ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=6)

            # ç”Ÿå¹´æœˆæ—¥
            ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=8)
            ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=10)
            style(ws.cell(row=4, column=11), border=thin_border)
        
            # æ°å
            ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
            ws.merge_cells(start_row=5, start_column=4, end_row=5, end_column=6)

            # æ€§åˆ¥
            ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=9)
            ws.merge_cells(start_row=5, start_column=10, end_row=5, end_column=TABLE_COLS)
        
            # ç¾ä½æ‰€
            ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)
            ws.merge_cells(start_row=6, start_column=4, end_row=6, end_column=6)

            # ç¨¼åƒå¯èƒ½æ—¥
            ws.merge_cells(start_row=6, start_column=7, end_row=6, end_column=9)
            ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=TABLE_COLS)
        
            # æœ€å¯„é§…
            ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
            ws.merge_cells(start_row=7, start_column=4, end_row=7, end_column=6)
            ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=TABLE_COLS)
            style(ws.cell(row=7, column=7), border=thin_border)

            # æœ€çµ‚å­¦æ­´
            ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
            ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=TABLE_COLS)
        
        
            # --- 9è¡Œç›®: 2. è³‡æ ¼ ---        
            qlist = [q.strip() for q in st.session_state.pi_qualifications_input.split("\n") if q.strip()]
            if not qlist: qlist = [""]
            
            for q in qlist:
                style(ws.cell(row=cur, column=2, value="æƒ…å ±å‡¦ç†è³‡æ ¼"), font=bold_font, border=thin_border)
                cell = ws.cell(row=cur, column=4, value=f"{q}")
                style(cell, border=thin_border)
            
                # è³‡æ ¼æ¬„ã¯ãƒ†ãƒ¼ãƒ–ãƒ«å¹…(Kåˆ—)ã¾ã§çµåˆ
                ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=3)
                ws.merge_cells(start_row=cur, start_column=4, end_row=cur, end_column=TABLE_COLS)
                cur += 1

            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=TABLE_COLS)
            style(ws.cell(row=cur, column=2), border=thin_border)
            cur += 1 # ç©ºç™½è¡Œ

            # --- 11è¡Œç›®: é–‹ç™ºçµŒé¨“ã‚µãƒãƒª ---
            cell = ws.cell(row=cur, column=2, value="é–‹ç™ºçµŒé¨“ã‚µãƒãƒª")
            style(cell, font=section_title_font, border=thin_border)
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=TABLE_COLS)
            cur += 1
        
            # ã‚µãƒãƒªæœ¬æ–‡ã‚‚ãƒ†ãƒ¼ãƒ–ãƒ«å¹…(Kåˆ—)ã¾ã§çµåˆ
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=TABLE_COLS)
            style(ws.cell(row=cur, column=2, value=st.session_state.pi_summary), border=thin_border, align=wrap_text_alignment)
            cur += 2 # ç©ºç™½è¡Œã‚’1ã¤æŒŸã‚€

            ws.merge_cells(start_row=cur - 1, start_column=2, end_row=cur - 1, end_column=TABLE_COLS)
            style(ws.cell(row=cur - 1, column=2), border=thin_border)
        
            # --- 17è¡Œç›®: 4. æ¥­å‹™çµŒæ­´ ---
            cell = ws.cell(row=cur, column=2, value="æ¥­å‹™çµŒæ­´")
            style(cell, font=section_title_font, fill=project_title_fill, border=thin_border)
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=TABLE_COLS)
            cur += 1
    
            # --- 18è¡Œç›®: æ¥­å‹™çµŒæ­´ãƒ†ãƒ¼ãƒ–ãƒ«ãƒ˜ãƒƒãƒ€ ---
            headers = [
                "é …ç•ª", "ä½œæ¥­æœŸé–“", "æ¡ˆä»¶å", "ä½œæ¥­å†…å®¹", "æ©Ÿç¨®", "è¨€èª/ãƒ„ãƒ¼ãƒ«", "ä½œæ¥­å·¥ç¨‹", "è¦æ¨¡",
                "æ¥­ç¨®", "OS", "DB/DC", "å½¹å‰²", "ãƒã‚¸ã‚·ãƒ§ãƒ³"
            ] # Båˆ—ã‹ã‚‰Kåˆ—

            targets = [
                (0,0), (0,1), (0,2), (0,3), (0,5), (0,6), (0,8), (0,9),
                (1,2), (1,5), (1,6), (1,8), (1,9)
            ]

            for i, (row, col) in enumerate(targets):
                if i < len(headers):
                    cell = ws.cell(row=cur + row, column=col + 2, value=headers[i])
                    style(cell, font=bold_font, fill=project_title_fill, border=thin_border, align=wrap_text_alignment)

            # ãƒ•ãƒªã‚¬ãƒŠ
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur + 1, end_column=2)
            ws.merge_cells(start_row=cur, start_column=3, end_row=cur + 1, end_column=3)
            ws.merge_cells(start_row=cur, start_column=5, end_row=cur + 1, end_column=6)
            ws.merge_cells(start_row=cur, start_column=8, end_row=cur, end_column=9)
            ws.merge_cells(start_row=cur + 1, start_column=8, end_row=cur + 1, end_column=9)
            
            cur += 2

            # --- 21è¡Œç›®ä»¥é™: æ¡ˆä»¶ãƒ«ãƒ¼ãƒ— (ãƒ†ãƒ¼ãƒ–ãƒ«å½¢å¼) ---
            for i, p in enumerate(st.session_state.projects):
                start_row = cur # ã“ã®æ¡ˆä»¶ã®é–‹å§‹è¡Œã‚’è¨˜æ†¶

                # 1åˆ—ç›®æ›¸ãè¾¼ã¿
                cell = ws.cell(row=start_row, column=2, value=i + 1)
                
                # 1åˆ—ç›®ã¯å…¨åˆ—ã«ç½«ç·šã¨æŠ˜ã‚Šè¿”ã—ã€ä¸Šå¯„ã›
                style(cell, font=work_history_font, border=thin_border, align=wrap_text_alignment)

                # --- 2è¡Œç›® (ä½œæ¥­æœŸé–“) ---
                start_date_str = p.get("start_date").strftime("%Y/%m/%d") if p.get("start_date") else ""
                end_date_str = p.get("end_date").strftime("%Y/%m/%d") if p.get("end_date") else ""
                delta_txt = ""
                if p.get("start_date") and p.get("end_date"):
                    days = (p["end_date"] - p["start_date"]).days
                    #delta_txt = f"(ç´„{round(days/30.4375,1)}ãƒ¶æœˆ)" if days >= 0 else "ï¼ˆ0ãƒ¶æœˆï¼‰"
                
                    if days >= 0:
                        # å…¨ä½“ã®æœˆæ•°ã‚’è¨ˆç®—
                        total_months = days / 30.4375
                    
                        # å¹´ã¨æœˆã«åˆ†è§£
                        years = int(total_months // 12)      # å¹´ï¼ˆæ•´æ•°ï¼‰
                        months = round(total_months % 12, 1) # ä½™ã‚Šã®æœˆï¼ˆå°æ•°ç¬¬1ä½ã¾ã§ï¼‰

                        # å››æ¨äº”å…¥ã§æœˆãŒã€Œ12.0ãƒ¶æœˆã€ã«ãªã£ãŸå ´åˆã¯ã€1å¹´ã«ç¹°ã‚Šä¸Šã’ã‚‹
                        if months == 12:
                            years += 1
                            months = 0

                        if years > 0:
                            # 1å¹´ä»¥ä¸Šã®å ´åˆï¼ˆä¾‹: ç´„2å¹´3.5ãƒ¶æœˆï¼‰
                            delta_txt = f"ç´„{years}å¹´{months}ãƒ¶æœˆ"
                        else:
                            # 1å¹´æœªæº€ã®å ´åˆï¼ˆä¾‹: ç´„6.5ãƒ¶æœˆï¼‰
                            delta_txt = f"ç´„{months}ãƒ¶æœˆ"
                    else:
                        delta_txt = "0ãƒ¶æœˆ"
                
                style(ws.cell(row=start_row, column=3, value=start_date_str),font=data_font, align=center_text_alignment, border=data_border)
                style(ws.cell(row=start_row + 1, column=3, value="ï½"),font=data_font, align=center_text_alignment, border=data_border)
                style(ws.cell(row=start_row + 2, column=3, value=end_date_str),font=data_font, align=center_text_alignment, border=data_border)
                style(ws.cell(row=start_row + 3, column=3, value=delta_txt),font=data_font, align=center_text_alignment, border=data_border)

                # --- 3è¡Œç›® (æ¡ˆä»¶åãƒ»æ¥­ç¨®) ---
                style(ws.cell(row=start_row, column=4, value=p.get("project_name","")), font=work_history_font)
                style(ws.cell(row=start_row + 1, column=4, value=p.get("industry","")), font=work_history_font)
            
                # --- 4è¡Œç›® (ä½œæ¥­å†…å®¹) ---
                content_lines = [line.strip() for line in str(p.get("work_content", "")).split("\n") if line.strip()]
                if not content_lines:
                    content_lines = [""]

                 # ç©ºã§ã‚‚4è¡Œã¯ç¢ºä¿
                if len(content_lines) < 4:
                    padding_needed = 4 - len(content_lines)
                    content_lines.extend([""] * padding_needed)
                    
                content_count = 0
                
                for line in content_lines:
                    # Cåˆ— (æ¡ˆä»¶åã®çœŸä¸‹) ã«ä½œæ¥­å†…å®¹ã‚’æ›¸ãè¾¼ã‚€
                    cell = ws.cell(row=cur, column=COL_PROJECT_NAME, value=line)
                    style(cell, border=dashdot_border, align=wrap_text_alignment)
                    
                    # ä½œæ¥­å†…å®¹ã‚»ãƒ«ã‚’æ¨ªã«çµåˆ (Cåˆ—ã‹ã‚‰Kåˆ—ã¾ã§)
                    ws.merge_cells(start_row=cur, start_column=COL_PROJECT_NAME, end_row=cur, end_column=COL_PROJECT_NAME + 1)
                    ws.merge_cells(start_row=cur, start_column=COL_PROJECT_NAME + 3, end_row=cur, end_column=COL_PROJECT_NAME + 4)
                
                    cur += 1 # æ¬¡ã®è¡Œã¸
                    content_count += 1

                # --- 7è¡Œç›® (æ©Ÿç¨®ãƒ»OS) ---
                os = [s.strip() for s in p.get("os", "").split("/") if s.strip()]
            
                for model in range(len(os)):
                    style(ws.cell(row=start_row + model, column=7, value=os[model]), font=work_history_font)
            
                # --- 8è¡Œç›® (è¨€èª/ãƒ„ãƒ¼ãƒ«ãƒ»DB/DC) ---
                lang_tool = [s.strip() for s in p.get("lang_tool", "").split("/") if s.strip()]
                db_dc = [s.strip() for s in p.get("db_dc", "").split("/") if s.strip()]
            
                lang_count = 0
                db_count = 0
            
                for lang in range(len(lang_tool)):
                    style(ws.cell(row=start_row + lang, column=8, value=lang_tool[lang]), font=work_history_font)
                    lang_count += 1

                if lang_tool != db_dc:
                    for db in range(len(db_dc)):
                        style(ws.cell(row=start_row + db + (lang_count + 1), column=8, value=db_dc[db]), font=work_history_font)
                        db_count += 1

                #st.write("å¤‰æ›´å‰:", cur, lang_count, db_count, content_count, lang_count + db_count - content_count, cur + lang_count + db_count - content_count)
            
                # --- 10è¡Œç›® (ä½œæ¥­å·¥ç¨‹ãƒ»å½¹å‰²) ---
                REVERSE_WORK_PROCESS_MAP = {v: k for k, v in WORK_PROCESS_MAP.items()}
                label_count = 0
            
                for j, label in enumerate(p.get("work_process_list", [])):
                    # é€†å¼•ããƒãƒƒãƒ—ã«å­˜åœ¨ã™ã‚‹ã‹ç¢ºèª
                    if label in REVERSE_WORK_PROCESS_MAP:
                        style(ws.cell(row=start_row + j, column=10, value=label), font=work_history_font)
                        label_count += 1

                style(ws.cell(row=start_row + label_count, column=10, value=p.get("role","")), font=work_history_font)
                label_count += 1

                # ç©ºã§ã‚‚4è¡Œã¯ç¢ºä¿
                if (lang_count + db_count - content_count) < 4:
                    # å€¤ãŒã€Œ-ã€ã ã£ãŸã‚‰
                    if (lang_count + db_count - content_count) < 0:
                        lang_count += (lang_count + db_count - content_count) * -1
                    else:
                        lang_count -= lang_count + db_count - content_count
                

                #st.write("å¤‰æ›´å¾Œ:", cur, lang_count, db_count, content_count, lang_count + db_count - content_count, cur + lang_count + db_count - content_count)

                if (lang_count + db_count - content_count) < (label_count - content_count):
                    cur += label_count - content_count
                else:
                    cur += lang_count + db_count - content_count
                
                # --- 11è¡Œç›® (è¦æ¨¡ãƒ»ãƒã‚¸ã‚·ãƒ§ãƒ³) ---
                style(ws.cell(row=start_row, column=TABLE_COLS, value=p.get("scale","")), font=work_history_font)
                style(ws.cell(row=start_row + 1, column=TABLE_COLS, value=p.get("position","")), font=work_history_font)                        
            
                # --- ã“ã®æ¡ˆä»¶ã®ç¸¦ã‚»ãƒ«çµåˆ ---
                end_row = cur - 1 # ã“ã®æ¡ˆä»¶ã®æœ€çµ‚è¡Œ
                if end_row > start_row: # ä½œæ¥­å†…å®¹ãªã©ã§2è¡Œä»¥ä¸Šã«ãªã£ãŸå ´åˆ
                    # Cåˆ— (æ¡ˆä»¶å/ä½œæ¥­å†…å®¹) ä»¥å¤–ã‚’ç¸¦ã«çµåˆ
                    for c_idx in [c for c in range(1, TABLE_COLS + 1) if c != COL_PROJECT_NAME]:
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                        # çµåˆã—ãŸã‚»ãƒ«ã®ã‚¹ã‚¿ã‚¤ãƒ«ã‚’å†é©ç”¨ (ä¸Šå¯„ã›)
                        cell = ws.cell(row=start_row, column=c_idx)
                        style(cell, align=wrap_text_alignment)
            
                for j in range((end_row + 1) - start_row):
                    # ä»–ã®åˆ— (A, B, D-K) ã«ã‚‚ç½«ç·šã‚’å¼•ã (çµåˆã•ã‚Œã‚‹è¦ªã‚»ãƒ«ä»¥å¤–)
                    for c_idx in [c for c in range(3, TABLE_COLS) if c != COL_PROJECT_NAME]:
                        style(ws.cell(row=start_row + j, column=c_idx + 1),font=work_history_font, border=dashdot_border)

                    style(ws.cell(row=start_row + j, column=2), border=Border(left=Side(style='thick'), right=Side(style='thin')))
                    style(ws.cell(row=start_row + j, column=TABLE_COLS), border=Border(right=Side(style='thick')))

                for j in range(8):
                    if j == 0:
                        style(ws.cell(row=start_row, column=3 + j), align=center_text_alignment, border=Border(top=Side(style='thick')))
                        style(ws.cell(row=end_row, column=3 + j), border=Border(bottom=Side(style='thick')))
                    else:
                        style(ws.cell(row=start_row, column=3 + j), border=Border(left=Side(style='dashDot'), right=Side(style='dashDot'), top=Side(style='thick')))
                        style(ws.cell(row=end_row, column=3 + j), border=Border(left=Side(style='dashDot'), right=Side(style='dashDot'), bottom=Side(style='thick')))
                
                style(ws.cell(row=start_row, column=2), border=Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick')))
                style(ws.cell(row=start_row, column=TABLE_COLS), border=Border(left=Side(style='dashDot'),right=Side(style='thick'), top=Side(style='thick')))
                style(ws.cell(row=end_row, column=2), border=Border(left=Side(style='thick'), bottom=Side(style='thick')))
                style(ws.cell(row=end_row, column=TABLE_COLS), border=Border(right=Side(style='thick'), bottom=Side(style='thick')))

            # --- å¹…èª¿æ•´ (ã‚µãƒ³ãƒ—ãƒ«å½¢å¼) ---
            ws.column_dimensions["A"].width = 1.3  # é …ç•ª
            ws.column_dimensions["B"].width = 3 # æœŸé–“
            ws.column_dimensions["C"].width = 13 # æ¡ˆä»¶å/ä½œæ¥­å†…å®¹
            ws.column_dimensions["D"].width = 15 # æ¥­ç¨®
            ws.column_dimensions["E"].width = 11.5 # OS
            ws.column_dimensions["F"].width = 20 # è¨€èª
            ws.column_dimensions["G"].width = 11.5 # DB
            ws.column_dimensions["H"].width = 4.25 # å·¥ç¨‹
            ws.column_dimensions["I"].width = 10.25 # å½¹å‰²
            ws.column_dimensions["J"].width = 22 # ãƒã‚¸ã‚·ãƒ§ãƒ³
            ws.column_dimensions["K"].width = 11 # è¦æ¨¡

            ws.row_dimensions[1].height = 43
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
    
        st.download_button(
            label="ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
            data=output.getvalue(),
            file_name=f"{st.session_state.pi_name or 'ã‚¹ã‚­ãƒ«ã‚·ãƒ¼ãƒˆ'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("Excelã‚’ç”Ÿæˆã—ã¾ã—ãŸã€‚")

if page == "ãƒ›ãƒ¼ãƒ ":
    basic_info()
    deve_expe()
    business_history()
    ai_impr()
elif page == "åŸºæœ¬æƒ…å ±":
    basic_info()
elif page == "é–‹ç™ºçµŒé¨“ã‚µãƒãƒª":
    deve_expe()
elif page == "æ¥­å‹™å±¥æ­´":
    business_history()
elif page == "AIã«ã‚ˆã‚‹æ”¹å–„":
    ai_impr()
